import os
import re
from html import unescape

import pandas as pd
import scrapy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from scrapy import signals


class ProfesiaItSpider(scrapy.Spider):

    name = "profesia_it"
    allowed_domains = ["profesia.sk"]
    start_urls = [
        "https://www.profesia.sk/praca/",
    ]

    custom_settings = {
        "FEED_FORMAT": "csv",
        "FEED_URI": "jobs.csv",
        "TELNETCONSOLE_ENABLED": False,
        "LOG_LEVEL": "INFO",
        "FEED_EXPORT_FIELDS": [
            "index",
            "title",
            "company",
            "location",
            "date_posted",
            "url",
            "salary",
            "description",
        ],
    }
    description_start_markers = (
        "čo budeš robiť",
        "co budes robit",
        "požiadavky na zamestnanca",
        "pozadavky na zamestnanca",
        "informácie o pracovnom mieste",
        "informacie o pracovnom mieste",
        "job description",
    )
    description_end_markers = (
        "inzerujúca spoločnosť",
        "inzerujuca spolocnost",
        "stručná charakteristika spoločnosti",
        "strucna charakteristika spolocnosti",
        "company profile",
        "kontakt",
        "contact",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.counter = 0

        try:
            with open("keywords.txt", "r", encoding="utf-8") as f:
                self.keywords = [line.strip().lower() for line in f if line.strip()]
        except FileNotFoundError:
            self.keywords = []
            self.logger.warning("keywords.txt not found – no rows will be highlighted.")

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super().from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_closed, signal=signals.spider_closed)
        return spider

    def parse(self, response):
        for job in response.css("ul.list li.list-row"):
            job_href = job.css("h2 a::attr(href)").get()
            if not job_href:
                continue
            job_url = response.urljoin(job_href)
            if not job_url or "/praca/" not in job_url or "/O" not in job_url:
                continue

            self.counter += 1

            salary_parts = job.css(
                'span.label-group a[data-dimension7="Salary label"] span.label::text'
            ).getall()
            salary = "".join(salary_parts).strip()

            item = {
                "index": self.counter,
                "title": job.css("span.title::text").get(default="").strip(),
                "company": job.css("span.employer::text").get(default="").strip(),
                "location": job.css("span.job-location::text").get(default="").strip(),
                "date_posted": job.css("div.list-footer span.info strong::text").get(
                    default=""
                ).strip(),
                "url": job_url,
                "salary": salary,
                # Keep schema stable while scraping all-site listings quickly.
                "description": "",
            }

            yield response.follow(
                job_url,
                callback=self.parse_job_detail,
                meta={"item": item},
                dont_filter=True,
            )

        # Handle pagination
        next_page = response.xpath("//link[@rel='next']/@href").get()
        if next_page:
            yield response.follow(next_page, callback=self.parse)

    @staticmethod
    def collapse_spaces(value: str) -> str:
        return re.sub(r"\s+", " ", unescape(value or "")).strip()

    def extract_description(self, response) -> str:
        text_nodes = response.css("main ::text").getall() or response.css("body ::text").getall()
        lines = []
        seen = set()
        for node in text_nodes:
            normalized = self.collapse_spaces(node)
            if not normalized:
                continue
            key = normalized.lower()
            if key in seen:
                continue
            seen.add(key)
            lines.append(normalized)

        if not lines:
            return ""

        lower_lines = [line.lower() for line in lines]
        start_idx = None
        for idx, line in enumerate(lower_lines):
            if any(marker in line for marker in self.description_start_markers):
                start_idx = idx
                break

        if start_idx is not None:
            lines = lines[start_idx:]
            lower_lines = lower_lines[start_idx:]
            for idx, line in enumerate(lower_lines):
                if idx == 0:
                    continue
                if any(marker in line for marker in self.description_end_markers):
                    lines = lines[:idx]
                    break

        description = "\n".join(lines)
        if len(description) > 12000:
            return description[:12000].rstrip()
        return description

    def parse_job_detail(self, response):
        item = dict(response.meta.get("item") or {})
        item["description"] = self.extract_description(response)
        yield item

    def spider_closed(self, spider):
        csv_path = "jobs.csv"
        xlsx_path = "jobs.xlsx"

        if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
            self.logger.warning("jobs.csv does not exist or is empty. Skipping XLSX export.")
            return

        # Read the CSV and export to XLSX
        df = pd.read_csv(csv_path, dtype=str)
        if df.empty:
            self.logger.warning("jobs.csv has no rows. Skipping XLSX export.")
            return
        df.to_excel(xlsx_path, index=False)

        wb = load_workbook(xlsx_path)
        ws = wb.active

        highlight = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            title_idx = headers.index("title") + 1
        except ValueError:
            self.logger.error("No 'title' column found.")
            wb.save(xlsx_path)
            return

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            title = str(row[title_idx - 1].value or "").lower()
            if any(kw in title for kw in self.keywords):
                for cell in row:
                    cell.fill = highlight

        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
            )
            max_length = min(max_length, 100)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(xlsx_path)
